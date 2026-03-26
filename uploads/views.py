from django.shortcuts import render
import csv
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.apps import apps
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from Employee.models import Employee


# -------------------------------------------------------------------
# Get model dynamically
# -------------------------------------------------------------------
def get_model_by_name(model_name):
    # Normalize model name
    normalized_name = model_name.lower().strip()
    
    # Handle common name variations
    name_mapping = {
        "matgattribute": "matgattributeitem",
        "materialattribute": "matgattributeitem",
        "matgattributeitem": "matgattributeitem",
    }
    
    # Check if we have a mapping
    if normalized_name in name_mapping:
        normalized_name = name_mapping[normalized_name]
    
    # Search for exact match first
    for model in apps.get_models():
        if model.__name__.lower() == normalized_name:
            return model
    
    # If not found, try partial match (e.g., "MatgAttribute" matches "MatgAttributeItem")
    for model in apps.get_models():
        if normalized_name in model.__name__.lower() or model.__name__.lower() in normalized_name:
            return model
    
    return None


# -------------------------------------------------------------------
# Convert CSV/JSON/Excel value → correct Python type
# -------------------------------------------------------------------
def convert_value(field, value):
    if value is None or value == "":
        return None

    internal_type = field.get_internal_type()

    try:
        if internal_type in ["IntegerField", "BigIntegerField"]:
            return int(value)

        if internal_type == "BooleanField":
            return str(value).lower() in ["1", "true", "yes"]

        if internal_type == "FloatField":
            return float(value)

        if internal_type == "DateField":
            return timezone.datetime.fromisoformat(value).date()

        if internal_type == "DateTimeField":
            return timezone.datetime.fromisoformat(value)

        if internal_type == "ForeignKey":
            # FKs must match to_field, not always "id"
            return field.related_model.objects.get(**{field.target_field.name: value})

        return value

    except Exception:
        return value  # safe fallback


# -------------------------------------------------------------------
# Handler: ItemMaster Phase 1 — Insert Base Data
# -------------------------------------------------------------------
def handle_itemmaster_phase_1(data, request):
    from itemmaster.models import ItemMaster
    from MaterialType.models import MaterialType
    from matgroups.models import MatGroup

    now = timezone.now()
    objs = []
    errors = []

    for idx, row in enumerate(data):
        try:
            # Helper function to get value by multiple possible keys (handles different header formats)
            def get_value(row, possible_keys):
                for key in possible_keys:
                    if key in row and row[key] and str(row[key]).strip():
                        return str(row[key]).strip()
                return None
            
            # Get mat_type_code (handle different header formats)
            mat_type_code_value = get_value(row, ["mat_type_code", "Mat Type Code", "mat type code", "MAT_TYPE_CODE"])
            mat_type_code = None
            if mat_type_code_value:
                mat_type_code = MaterialType.objects.filter(mat_type_code=mat_type_code_value).first()
                if not mat_type_code:
                    errors.append({"row": idx + 2, "error": f"MaterialType '{mat_type_code_value}' not found"})
                    continue
            else:
                errors.append({"row": idx + 2, "error": "mat_type_code is required"})
                continue
            
            # Get mgrp_code (handle different header formats)
            mgrp_code_value = get_value(row, ["mgrp_code", "Mgrp Code", "mgrp code", "MGRP_CODE"])
            mgrp_code = None
            if mgrp_code_value:
                mgrp_code = MatGroup.objects.filter(mgrp_code=mgrp_code_value).first()
                if not mgrp_code:
                    errors.append({"row": idx + 2, "error": f"MatGroup '{mgrp_code_value}' not found"})
                    continue
            else:
                errors.append({"row": idx + 2, "error": "mgrp_code is required"})
                continue
            
            # Get short_name (handle different header formats)
            short_name = get_value(row, ["short_name", "Short Name", "short name", "SHORT_NAME"]) or ""
            if not short_name:
                errors.append({"row": idx + 2, "error": "short_name is required"})
                continue
            
            # Convert sap_item_id to int if it's a string
            sap_item_id_value = get_value(row, ["sap_item_id", "Sap Item Id", "sap item id", "SAP_ITEM_ID"])
            sap_item_id = None
            if sap_item_id_value:
                try:
                    sap_item_id = int(float(sap_item_id_value))
                except (ValueError, TypeError):
                    sap_item_id = None
            
            # Get other optional fields
            long_name = get_value(row, ["long_name", "Long Name", "long name", "LONG_NAME"])
            mgrp_long_name = get_value(row, ["mgrp_long_name", "Mgrp Long Name", "mgrp long name", "MGRP_LONG_NAME"])
            sap_name = get_value(row, ["sap_name", "Sap Name", "sap name", "SAP_NAME"])
            search_text = get_value(row, ["search_text", "Search Text", "search text", "SEARCH_TEXT"])
            
            objs.append(ItemMaster(
                sap_item_id=sap_item_id,
                mat_type_code=mat_type_code,
                mgrp_code=mgrp_code,
                short_name=short_name,
                long_name=long_name,
                mgrp_long_name=mgrp_long_name,
                sap_name=sap_name,
                search_text=search_text,
                created=now,
                updated=now,
            ))
        except Exception as e:
            import traceback
            errors.append({"row": idx + 2, "error": f"{str(e)}"})

    if objs:
        ItemMaster.objects.bulk_create(objs, ignore_conflicts=True)

    return JsonResponse({
        "message": "ItemMaster Phase 1 upload complete",
        "inserted": len(objs),
        "errors": errors,
    })


# -------------------------------------------------------------------
# Validation Functions
# -------------------------------------------------------------------
def validate_attribute_value(value, validation_type):
    """
    Validate attribute value based on validation type.
    Returns (is_valid, error_message)
    """
    if not validation_type or not value:
        return True, None
    
    validation_type = validation_type.lower().strip()
    
    if validation_type == "alpha":
        if not value.replace(" ", "").isalpha():
            return False, f"Value '{value}' must contain only alphabetic characters"
    
    elif validation_type == "numeric":
        if not value.replace(".", "").replace("-", "").isdigit():
            return False, f"Value '{value}' must be numeric"
    
    elif validation_type == "alphanumeric":
        if not value.replace(" ", "").isalnum():
            return False, f"Value '{value}' must contain only alphanumeric characters"
    
    elif validation_type == "wholenumber":
        try:
            num = float(value)
            if num < 0 or num != int(num):
                return False, f"Value '{value}' must be a whole number (non-negative integer)"
        except ValueError:
            return False, f"Value '{value}' must be a whole number"
    
    elif validation_type == "integer":
        try:
            int(value)
        except ValueError:
            return False, f"Value '{value}' must be an integer"
    
    elif validation_type == "decimal":
        try:
            float(value)
        except ValueError:
            return False, f"Value '{value}' must be a decimal number"
    
    return True, None


# -------------------------------------------------------------------
# Handler: ItemMaster Phase 2 — Merge Attributes JSON
# -------------------------------------------------------------------
def handle_itemmaster_phase_2(data, request):
    from itemmaster.models import ItemMaster
    from matg_attributes.models import MatgAttributeItem
    import json

    updated = 0
    created = 0
    unchanged = 0
    errors = []

    # Helper function to get value by multiple possible keys (handles different header formats)
    def get_value(row, possible_keys):
        for key in possible_keys:
            if key in row and row[key] is not None and str(row[key]).strip():
                return str(row[key]).strip()
        return None

    for idx, row in enumerate(data, start=2):  # start=2 because row 1 is header
        try:
            # Get sap_item_id (handle different header formats)
            sap_item_id_value = get_value(row, ["sap_item_id", "Sap Item Id", "sap item id", "SAP_ITEM_ID"])
            if not sap_item_id_value:
                errors.append({"row": idx, "error": "sap_item_id is required"})
                continue
            
            # Convert to int
            try:
                sap = int(float(sap_item_id_value))
            except (ValueError, TypeError):
                errors.append({"row": idx, "error": f"Invalid sap_item_id: {sap_item_id_value}"})
                continue

            # Find item
            item = ItemMaster.objects.filter(sap_item_id=sap).first()
            if not item:
                errors.append({"row": idx, "error": f"ItemMaster with sap_item_id {sap} not found"})
                continue

            # Get attribute_name (handle different header formats)
            attr_name = get_value(row, ["attribute_name", "Attribute Name", "attribute name", "ATTRIBUTE_NAME"])
            if not attr_name:
                errors.append({"row": idx, "error": "attribute_name is required"})
                continue

            # Get attribute_value (handle different header formats)
            attr_value = get_value(row, ["attribute_value", "Attribute Value", "attribute value", "ATTRIBUTE_VALUE"])
            if attr_value is None:  # Allow empty string but not None
                attr_value = ""

            # Get UOM (handle different header formats) - optional
            uom = get_value(row, ["uom", "Uom", "UOM", "Unit Of Measure", "unit of measure"])

            # Look up the attribute definition to get validation rules
            attr_def = None
            try:
                attr_def = MatgAttributeItem.objects.filter(
                    mgrp_code=item.mgrp_code,
                    attribute_name=attr_name,
                    is_deleted=False
                ).first()
            except Exception as e:
                # If we can't find the attribute definition, we'll skip validation
                pass

            # Validate attribute value if validation rule exists
            if attr_def and attr_def.validation and attr_value:
                is_valid, error_msg = validate_attribute_value(attr_value, attr_def.validation)
                if not is_valid:
                    errors.append({"row": idx, "error": error_msg})
                    continue
                # If validation passes, allow the value even if it's not in possible_values (custom values are allowed)
            # If no validation type is set, check possible_values
            elif attr_def and attr_def.possible_values and len(attr_def.possible_values) > 0:
                if attr_value not in attr_def.possible_values:
                    errors.append({
                        "row": idx,
                        "error": f"Value '{attr_value}' is not in allowed values: {', '.join(attr_def.possible_values)}"
                    })
                    continue

            # Ensure JSON is dict
            attributes = item.attributes or {}
            if isinstance(attributes, str):
                try:
                    attributes = json.loads(attributes)
                except json.JSONDecodeError:
                    attributes = {}

            # Get old value before updating (for tracking changes)
            old_attr_data = attributes.get(attr_name)
            old_value = None
            if isinstance(old_attr_data, dict):
                old_value = old_attr_data.get("value", "")
            elif old_attr_data is not None:
                old_value = str(old_attr_data)

            # Store attribute value (with UOM if provided)
            # Store as: {"AttributeName": "value"} or {"AttributeName": {"value": "value", "uom": "kg"}}
            if uom:
                attributes[attr_name] = {"value": attr_value, "uom": uom}
            else:
                attributes[attr_name] = attr_value

            # Track changes
            if old_value is None or old_value == "":
                created += 1
            elif old_value != attr_value:
                updated += 1
            else:
                unchanged += 1

            # Update item attributes
            item.attributes = attributes
            item.save()

        except Exception as e:
            import traceback
            errors.append({"row": idx, "error": f"{str(e)}"})

    return JsonResponse({
        "message": "ItemMaster Phase 2 attribute merge complete",
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "errors": errors,
    })


# -------------------------------------------------------------------
# Generic Handler: For any model that doesn't have a specific handler
# -------------------------------------------------------------------
def handle_generic_model_upload(data, request, Model, model_name):
    """
    Generic handler for bulk uploading any model.
    Automatically handles ForeignKey fields and converts data types.
    """
    now = timezone.now()
    objs = []
    errors = []
    
    # Get the user from request if available (for audit fields)
    user = None
    # Try to get user from token if available
    try:
        from Employee.decorator import get_user_from_token
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if token:
            user = get_user_from_token(token)
    except:
        pass
    
    # Get all model fields (only concrete fields, not reverse relations)
    model_fields = {}
    for field in Model._meta.concrete_fields:
        model_fields[field.name] = field
    
    for idx, row in enumerate(data, start=2):  # start=2 because row 1 is header
        try:
            obj_data = {}
            
            # Process each field in the row
            for field_name, value in row.items():
                # Normalize field name (handle spaces, case differences)
                normalized_field_name = field_name.lower().replace(' ', '_').replace('-', '_').strip()
                
                # Find matching field (case-insensitive)
                matching_field = None
                for model_field_name, model_field in model_fields.items():
                    if model_field_name.lower() == normalized_field_name:
                        matching_field = model_field
                        break
                
                if not matching_field:
                    # Field not found in model, skip it (might be a header formatting issue)
                    continue
                
                # Skip audit fields (they'll be set automatically)
                if matching_field.name in ['id', 'created', 'updated', 'createdby', 'updatedby', 'is_deleted']:
                    continue
                
                # Handle ForeignKey fields
                if hasattr(matching_field, 'related_model') and matching_field.related_model:
                    if value and str(value).strip():
                        # Get the related model
                        related_model = matching_field.related_model
                        # Try to find by primary key or by the target field
                        target_field = matching_field.target_field.name if hasattr(matching_field, 'target_field') else 'pk'
                        
                        try:
                            # Try to get by the target field value
                            fk_obj = related_model.objects.filter(**{target_field: str(value).strip()}).first()
                            if fk_obj:
                                obj_data[matching_field.name] = fk_obj
                            else:
                                errors.append({
                                    "row": idx,
                                    "field": field_name,
                                    "error": f"Foreign key value '{value}' not found in {related_model.__name__}"
                                })
                        except Exception as e:
                            errors.append({
                                "row": idx,
                                "field": field_name,
                                "error": f"Error resolving foreign key: {str(e)}"
                            })
                else:
                    # Handle regular fields
                    try:
                        converted_value = convert_value(matching_field, value)
                        obj_data[matching_field.name] = converted_value
                    except Exception as e:
                        errors.append({
                            "row": idx,
                            "field": field_name,
                            "error": f"Error converting value: {str(e)}"
                        })
            
            # Set audit fields if they exist
            if 'created' in model_fields:
                obj_data['created'] = now
            if 'updated' in model_fields:
                obj_data['updated'] = now
            if 'createdby' in model_fields and user:
                obj_data['createdby'] = user
            if 'updatedby' in model_fields and user:
                obj_data['updatedby'] = user
            
            # Create the object
            if obj_data:
                obj = Model(**obj_data)
                objs.append(obj)
        
        except Exception as e:
            errors.append({
                "row": idx,
                "error": f"Error processing row: {str(e)}"
            })
    
    # Bulk create objects
    if objs:
        try:
            Model.objects.bulk_create(objs, ignore_conflicts=True)
        except Exception as e:
            return JsonResponse({
                "error": f"Bulk create failed: {str(e)}",
                "errors": errors
            }, status=400)
    
    return JsonResponse({
        "message": f"{model_name} upload complete",
        "inserted": len(objs),
        "errors": errors,
    })


# -------------------------------------------------------------------
# Handler: MatgAttributeItem Phase 1 — Insert Allowed Values + UOMs
# -------------------------------------------------------------------
def handle_matgattribute_phase_1(data, request):
    from matg_attributes.models import MatgAttributeItem
    from matgroups.models import MatGroup

    objs = []
    errors = []

    # Helper function to get value by multiple possible keys (handles different header formats)
    def get_value(row, possible_keys):
        for key in possible_keys:
            if key in row and row[key] and str(row[key]).strip():
                return str(row[key]).strip()
        return None

    for idx, row in enumerate(data, start=2):  # start=2 because row 1 is header
        try:
            # Get mgrp_code (handle different header formats)
            mgrp_code_value = get_value(row, ["mgrp_code", "Mgrp Code", "mgrp code", "MGRP_CODE"])
            mgrp_code = None
            if mgrp_code_value:
                mgrp_code = MatGroup.objects.filter(mgrp_code=mgrp_code_value).first()
                if not mgrp_code:
                    errors.append({"row": idx, "error": f"MatGroup '{mgrp_code_value}' not found"})
                    continue
            else:
                errors.append({"row": idx, "error": "mgrp_code is required"})
                continue
            
            # Get attribute_name (handle different header formats)
            attribute_name = get_value(row, ["attribute_name", "Attribute Name", "attribute name", "ATTRIBUTE_NAME"])
            if not attribute_name:
                errors.append({"row": idx, "error": "attribute_name is required"})
                continue
            
            # Get possible_values (handle different header formats)
            possible_values_str = get_value(row, ["possible_values", "Possible Values", "possible values", "POSSIBLE_VALUES"])
            possible_vals = []
            if possible_values_str:
                possible_vals = [
                    x.strip() for x in possible_values_str.split(",")
                    if x.strip()
                ]
            
            # Get uom (handle different header formats)
            uom_str = get_value(row, ["uom", "Uom", "UOM", "Unit Of Measure"])
            uom = uom_str if uom_str else None
            
            # Get print_priority (handle different header formats)
            print_priority_str = get_value(row, ["print_priority", "Print Priority", "print priority", "PRINT_PRIORITY"])
            print_priority = None
            if print_priority_str:
                try:
                    print_priority = int(float(print_priority_str))
                except (ValueError, TypeError):
                    print_priority = None
            
            # Get validation (handle different header formats)
            validation = get_value(row, ["validation", "Validation", "VALIDATION"])
            
            objs.append(MatgAttributeItem(
                mgrp_code=mgrp_code,
                attribute_name=attribute_name,
                possible_values=possible_vals,
                uom=uom,
                print_priority=print_priority,
                validation=validation,
            ))

        except Exception as e:
            import traceback
            errors.append({"row": idx, "error": f"{str(e)}"})

    if objs:
        try:
            MatgAttributeItem.objects.bulk_create(objs, ignore_conflicts=True)
        except Exception as e:
            return JsonResponse({
                "error": f"Bulk create failed: {str(e)}",
                "errors": errors
            }, status=400)

    return JsonResponse({
        "message": "MatGroup Attribute Definitions imported",
        "inserted": len(objs),
        "errors": errors,
    })


# -------------------------------------------------------------------
# MAIN BULK UPLOAD FUNCTION WITH PHASE ROUTING
# -------------------------------------------------------------------
@csrf_exempt
def bulk_upload(request):
    model_name = request.POST.get("model")
    phase = request.POST.get("phase", "1")

    if not model_name:
        return JsonResponse({"error": "Model name is required"}, status=400)

    Model = get_model_by_name(model_name)
    if not Model:
        return JsonResponse({"error": f"Invalid model: {model_name}"}, status=400)

    # -------------------------------------------------------------------
    # Parse file (CSV or Excel)
    # -------------------------------------------------------------------
    data = []
    file = request.FILES.get("file")

    if not file:
        return JsonResponse({"error": "No file uploaded"}, status=400)

    ext = file.name.split('.')[-1].lower()

    try:
        # Excel Upload
        if ext in ["xlsx", "xls"]:
            wb = openpyxl.load_workbook(file)
            
            # For ItemMaster phase 2, look for "Attributes" sheet
            if model_name and model_name.lower() == "itemmaster" and phase == "2":
                # Try to find Attributes sheet (case-insensitive)
                attributes_sheet = None
                for sheet_name in wb.sheetnames:
                    if sheet_name.lower().strip() == "attributes":
                        attributes_sheet = wb[sheet_name]
                        break
                
                if attributes_sheet:
                    sheet = attributes_sheet
                else:
                    # Fall back to active sheet if Attributes not found
                    sheet = wb.active
            else:
                # For phase 1 or other models, use active sheet
                sheet = wb.active
            
            header = [str(c.value).strip() if c.value else "" for c in next(sheet.rows)]  # Filter out None headers and convert to string
            
            # Remove empty headers
            header = [h for h in header if h]

            for row in sheet.iter_rows(min_row=2):
                row_dict = {}
                for idx, cell in enumerate(row):
                    if idx < len(header) and header[idx]:
                        cell_value = "" if cell.value is None else str(cell.value).strip()
                        row_dict[header[idx]] = cell_value
                # Only add non-empty rows (at least one non-empty value)
                if any(v for v in row_dict.values() if v):
                    data.append(row_dict)

        # CSV Upload
        elif ext == "csv":
            file_content = file.read().decode("utf-8")
            rows = csv.DictReader(file_content.splitlines())
            data = list(rows)

        else:
            return JsonResponse({"error": "Only CSV or Excel allowed"}, status=400)

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"File parsing error: {str(e)}")
        print(error_details)
        return JsonResponse({"error": f"File parsing failed: {str(e)}"}, status=400)

    if not data:
        return JsonResponse({"error": "File is empty"}, status=400)

    # -------------------------------------------------------------------
    # Route to model-phase handlers (case-insensitive)
    # -------------------------------------------------------------------
    model_name_lower = model_name.lower()
    
    if model_name_lower == "itemmaster":
        if phase == "1":
            return handle_itemmaster_phase_1(data, request)
        if phase == "2":
            return handle_itemmaster_phase_2(data, request)
        return JsonResponse({"error": f"Invalid phase '{phase}' for ItemMaster. Use phase=1 or phase=2"}, status=400)

    if model_name_lower == "matgattributeitem" or model_name == "MatgAttributeItem":
        if phase == "1":
            return handle_matgattribute_phase_1(data, request)
        return JsonResponse({"error": f"Invalid phase '{phase}' for MatgAttributeItem"}, status=400)

    # Generic handler for all other models
    return handle_generic_model_upload(data, request, Model, model_name)


def get_model_fields(request):
    model_name = request.GET.get("model")
    Model = get_model_by_name(model_name)

    if not Model:
        return JsonResponse({"error": "Invalid model"}, status=400)

    fields = [f.name for f in Model._meta.fields if f.name != "id"]
    return JsonResponse({"fields": fields})


# -------------------------------------------------------------------
# Generate ItemMaster Base Values Template
# -------------------------------------------------------------------
def generate_itemmaster_base_template(Model):
    """Generate template for ItemMaster base values (excluding attributes and is_final)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Item Base Values"
    
    # Fields to exclude
    exclude_fields = {'id', 'created', 'updated', 'createdby', 'updatedby', 'is_deleted', 
                     'createdby_id', 'updatedby_id', 'local_item_id', 'attributes', 'is_final'}
    
    # Get data entry fields
    data_entry_fields = []
    for field in Model._meta.concrete_fields:
        if field.name not in exclude_fields and not field.many_to_many:
            data_entry_fields.append(field)
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Create headers
    headers = []
    for field in data_entry_fields:
        if hasattr(field, 'db_column') and field.db_column:
            field_display = field.db_column.replace('_', ' ').title()
        else:
            field_display = field.name.replace('_', ' ').title()
        headers.append(field_display)
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25
    
    # Add sample data rows
    sample_data = [
        {"sap_item_id": "12345", "mat_type_code": "MAT1", "mgrp_code": "GRP001", "short_name": "Sample Item 1", "long_name": "Sample Long Name 1", "mgrp_long_name": "Material Group Long Name", "sap_name": "SAP Item Name 1", "search_text": "sample search text 1"},
        {"sap_item_id": "12346", "mat_type_code": "MAT2", "mgrp_code": "GRP002", "short_name": "Sample Item 2", "long_name": "Sample Long Name 2", "mgrp_long_name": "Material Group Long Name 2", "sap_name": "SAP Item Name 2", "search_text": "sample search text 2"},
    ]
    
    for row_idx, sample_row in enumerate(sample_data, start=2):
        for col_idx, field in enumerate(data_entry_fields, start=1):
            field_name = field.db_column if hasattr(field, 'db_column') and field.db_column else field.name
            value = sample_row.get(field_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "ItemMaster_Base_Values_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -------------------------------------------------------------------
# Generate MatgAttributeItem Template
# -------------------------------------------------------------------
def generate_matgattribute_template(Model):
    """Generate template for MatgAttributeItem with proper sample data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MatgAttributeItem"
    
    # Fields to exclude
    exclude_fields = {'id', 'created', 'updated', 'createdby', 'updatedby', 'is_deleted', 
                     'createdby_id', 'updatedby_id'}
    
    # Get data entry fields
    data_entry_fields = []
    for field in Model._meta.concrete_fields:
        if field.name not in exclude_fields and not field.many_to_many:
            data_entry_fields.append(field)
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Create headers
    headers = []
    for field in data_entry_fields:
        if hasattr(field, 'db_column') and field.db_column:
            field_display = field.db_column.replace('_', ' ').title()
        else:
            field_display = field.name.replace('_', ' ').title()
        headers.append(field_display)
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
    
    # Add sample data specific to MatgAttributeItem
    sample_data = [
        {
            "mgrp_code": "GRP001",
            "attribute_name": "Color",
            "possible_values": "Red, Blue, Green, Yellow",
            "uom": "",
            "print_priority": "1",
            "validation": ""
        },
        {
            "mgrp_code": "GRP001",
            "attribute_name": "Size",
            "possible_values": "Small, Medium, Large, XL",
            "uom": "",
            "print_priority": "2",
            "validation": ""
        },
        {
            "mgrp_code": "GRP002",
            "attribute_name": "Weight",
            "possible_values": "1kg, 2kg, 5kg, 10kg",
            "uom": "kg",
            "print_priority": "1",
            "validation": ""
        },
    ]
    
    for row_idx, sample_row in enumerate(sample_data, start=2):
        for col_idx, field in enumerate(data_entry_fields, start=1):
            field_name = field.db_column if hasattr(field, 'db_column') and field.db_column else field.name
            value = sample_row.get(field_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "MatgAttributeItem_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -------------------------------------------------------------------
# Generate ItemMaster Attributes Template
# -------------------------------------------------------------------
def generate_itemmaster_attributes_template():
    """Generate template for ItemMaster attributes"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attribute Settings"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Headers (including UOM)
    headers = ["Sap Item Id", "Attribute Name", "Attribute Value", "Uom"]
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30
    
    # Add sample data (including UOM)
    sample_data = [
        {"sap_item_id": "12345", "attribute_name": "Color", "attribute_value": "Red", "uom": ""},
        {"sap_item_id": "12345", "attribute_name": "Size", "attribute_value": "Large", "uom": ""},
        {"sap_item_id": "12346", "attribute_name": "Color", "attribute_value": "Blue", "uom": ""},
        {"sap_item_id": "12346", "attribute_name": "Weight", "attribute_value": "10", "uom": "kg"},
    ]
    
    for row_idx, sample_row in enumerate(sample_data, start=2):
        ws.cell(row=row_idx, column=1, value=sample_row["sap_item_id"])
        ws.cell(row=row_idx, column=2, value=sample_row["attribute_name"])
        ws.cell(row=row_idx, column=3, value=sample_row["attribute_value"])
        ws.cell(row=row_idx, column=4, value=sample_row["uom"])
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "ItemMaster_Attributes_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# -------------------------------------------------------------------
# Generate Excel Template Dynamically
# -------------------------------------------------------------------
@csrf_exempt
def generate_excel_template(request):
    model_name = request.GET.get("model")
    template_type = request.GET.get("type", "base")  # "base" or "attributes" for ItemMaster
    
    if not model_name:
        return JsonResponse({"error": "Model name is required"}, status=400)
    
    Model = get_model_by_name(model_name)
    if not Model:
        # Try to get available models for debugging
        try:
            available_models = [m.__name__ for m in apps.get_models()]
            return JsonResponse({
                "error": f"Invalid model: {model_name}",
                "available_models": sorted(available_models)
            }, status=400)
        except:
            return JsonResponse({"error": f"Invalid model: {model_name}"}, status=400)
    
    # For ItemMaster, handle separate downloads
    if model_name.lower() == "itemmaster":
        if template_type == "attributes":
            return generate_itemmaster_attributes_template()
        else:
            return generate_itemmaster_base_template(Model)
    
    # Special handling for MatgAttributeItem - better sample data
    if Model.__name__.lower() == "matgattributeitem":
        return generate_matgattribute_template(Model)
    
    # Fields to exclude (audit fields)
    exclude_fields = {'id', 'created', 'updated', 'createdby', 'updatedby', 'is_deleted', 
                     'createdby_id', 'updatedby_id', 'local_item_id'}
    
    # Get all fields from the model
    all_fields = Model._meta.concrete_fields
    
    # Filter fields - only include data entry fields
    data_entry_fields = []
    for field in all_fields:
        if field.name not in exclude_fields and not field.many_to_many:
            data_entry_fields.append(field)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = model_name
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Create headers
    headers = []
    for field in data_entry_fields:
        if hasattr(field, 'db_column') and field.db_column:
            field_display = field.db_column.replace('_', ' ').title()
        else:
            field_display = field.name.replace('_', ' ').title()
        headers.append(field_display)
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25
    
    # Add sample data based on field types
    sample_rows = []
    for i in range(2):  # Add 2 sample rows
        sample_row = {}
        for field in data_entry_fields:
            field_name = field.db_column if hasattr(field, 'db_column') and field.db_column else field.name
            field_type = field.get_internal_type()
            
            if field_type == 'ForeignKey':
                sample_row[field_name] = f"SAMPLE_FK_{i+1}"
            elif field_type == 'IntegerField':
                sample_row[field_name] = 1000 + i
            elif field_type == 'BooleanField':
                sample_row[field_name] = "True" if i == 0 else "False"
            elif field_type == 'CharField':
                max_length = field.max_length if hasattr(field, 'max_length') else 50
                sample_row[field_name] = f"Sample {field_name.replace('_', ' ').title()} {i+1}"[:max_length]
            elif field_type == 'JSONField':
                # For JSONField, provide comma-separated values as string
                sample_row[field_name] = "Value1, Value2, Value3"
            else:
                sample_row[field_name] = f"Sample Value {i+1}"
        
        sample_rows.append(sample_row)
    
    # Write sample data
    for row_idx, sample_row in enumerate(sample_rows, start=2):
        for col_idx, field in enumerate(data_entry_fields, start=1):
            field_name = field.db_column if hasattr(field, 'db_column') and field.db_column else field.name
            value = sample_row.get(field_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{model_name}_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{model_name}_template.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
